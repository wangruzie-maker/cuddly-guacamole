"""Export WeChat Channels accumulated results to Excel with embedded cover thumbnails."""

from __future__ import annotations

import io
import tempfile
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

CHANNELS_REFERER = "https://channels.weixin.qq.com/"

TEXT_COLUMNS = [
    ("链接", "url"),
    ("feed_id", "feed_id"),
    ("状态", "status"),
    ("类型", "note_type"),
    ("标题", "title"),
    ("文案", "desc"),
    ("视频脚本", "video_script"),
    ("脚本来源", "video_script_source"),
    ("脚本状态", "video_script_status"),
    ("作者", "author"),
    ("点赞数", "liked_count"),
    ("评论数", "comment_count"),
    ("分享数", "share_count"),
    ("收藏数", "collect_count"),
    ("位置", "location"),
    ("提取模式", "extract_mode"),
    ("视频URL", "video_url"),
    ("提取时间", "extracted_at"),
    ("错误信息", "error"),
]

COVER_COLUMN = "封面"
THUMB_SIZE = (140, 140)

DOWNLOAD_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Referer": CHANNELS_REFERER,
}


def _load_cover_path(url: str, feed_id: str, cache_dir: Path) -> Path | None:
    if not url:
        return None
    dest = cache_dir / feed_id / "cover.jpg"
    if dest.exists():
        return dest
    try:
        resp = requests.get(url, headers=DOWNLOAD_HEADERS, timeout=30)
        resp.raise_for_status()
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(resp.content)
        return dest
    except Exception:
        return None


def _make_thumbnail(src: Path, dest: Path) -> Path:
    with PILImage.open(src) as img:
        img = img.convert("RGB")
        img.thumbnail(THUMB_SIZE, PILImage.Resampling.LANCZOS)
        dest.parent.mkdir(parents=True, exist_ok=True)
        img.save(dest, format="JPEG", quality=85)
    return dest


def build_channels_excel_bytes(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "视频号"

    headers = [name for name, _ in TEXT_COLUMNS] + [COVER_COLUMN]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    cache_dir = Path(tempfile.gettempdir()) / "channels_excel_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    thumb_dir = cache_dir / "thumbs"
    cover_col = len(TEXT_COLUMNS) + 1
    cover_letter = get_column_letter(cover_col)
    ws.column_dimensions[cover_letter].width = 22

    for row_idx, row in enumerate(rows, start=2):
        values = [row.get(key, "") for _, key in TEXT_COLUMNS]
        values.append("")
        ws.append(values)
        ws.row_dimensions[row_idx].height = 110

        for col_idx, (_, key) in enumerate(TEXT_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

        feed_id = row.get("feed_id") or f"row{row_idx}"
        cover_url = row.get("cover_url") or ""
        src_path = _load_cover_path(cover_url, feed_id, cache_dir)
        if not src_path or not src_path.exists():
            continue

        thumb = thumb_dir / f"{feed_id}_{row_idx}.jpg"
        try:
            _make_thumbnail(src_path, thumb)
            xl_img = XLImage(str(thumb))
            xl_img.width, xl_img.height = THUMB_SIZE
            ws.add_image(xl_img, f"{cover_letter}{row_idx}")
        except Exception as exc:
            ws.cell(row=row_idx, column=cover_col, value=f"[封面加载失败: {exc}]")

    for col_idx in range(1, len(TEXT_COLUMNS) + 1):
        letter = get_column_letter(col_idx)
        key = TEXT_COLUMNS[col_idx - 1][1]
        width = 18 if key in {"url", "desc", "video_script", "video_url", "title"} else 12
        ws.column_dimensions[letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
