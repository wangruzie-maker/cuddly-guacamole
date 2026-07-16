"""Export watch-topic items to Excel."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

TEXT_COLUMNS: list[tuple[str, str]] = [
    ("标题", "title"),
    ("作者", "author"),
    ("内容类型", "note_type"),
    ("平台", "platform"),
    ("链接", "url"),
    ("搜索词", "keyword"),
    ("点赞", "liked_count"),
    ("收藏", "collected_count"),
    ("评论", "comment_count"),
    ("分享", "share_count"),
    ("浏览", "view_count"),
    ("转录类型", "_transcription_kind"),
    ("转录状态", "_transcription_status"),
    ("转录内容", "_transcription_text"),
]


def _row_value(item: dict[str, Any], key: str) -> Any:
    if key == "_transcription_kind":
        return (item.get("transcription") or {}).get("kind") or ""
    if key == "_transcription_status":
        transcription = item.get("transcription") or {}
        progress = transcription.get("progress") or {}
        if progress.get("label"):
            return progress["label"]
        status = transcription.get("status") or ""
        if status == "completed":
            return "已完成"
        if status == "running":
            return "转录中"
        if status == "failed":
            return "失败"
        return status or "未转录"
    if key == "_transcription_text":
        transcription = item.get("transcription") or {}
        text = str(transcription.get("text") or "")
        if transcription.get("truncated"):
            text += "…"
        return text
    return item.get(key, "")


def build_topic_items_excel_bytes(topic_name: str, items: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "采集结果"
    ws.append([label for label, _ in TEXT_COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for item in items:
        ws.append([_row_value(item, key) for _, key in TEXT_COLUMNS])

    for index, (label, key) in enumerate(TEXT_COLUMNS, start=1):
        width = 42 if key in {"url", "_transcription_text", "title"} else 14
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "A2"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
