"""Export accumulated results to Excel with embedded images."""

from __future__ import annotations

import io
import tempfile
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from fetch_extractor import DEFAULT_HEADERS
from image_cache import cache_note_images
from image_urls import dedupe_item_images

TEXT_COLUMNS = [
    ("链接", "url"),
    ("笔记ID", "feed_id"),
    ("状态", "status"),
    ("类型", "note_type"),
    ("标题", "title"),
    ("文案", "desc"),
    ("图片OCR", "image_ocr_text"),
    ("视频脚本", "video_script"),
    ("脚本来源", "video_script_source"),
    ("作者", "author"),
    ("点赞数", "liked_count"),
    ("收藏数", "collected_count"),
    ("评论数", "comment_count"),
    ("提取时间", "extracted_at"),
]

MAX_IMAGES_CAP = 30
THUMB_SIZE = (140, 140)


def _load_image_path(url: str, feed_id: str, index: int, cache_dir: Path) -> Path | None:
    local_paths = cache_dir / feed_id
    if local_paths.exists():
        candidates = sorted(local_paths.glob(f"img_{index:02d}.*"))
        if candidates:
            return candidates[0]

    try:
        saved = cache_note_images(feed_id, [url])
        if saved:
            return Path(saved[0])
    except Exception:
        pass

    try:
        resp = requests.get(url, headers=DEFAULT_HEADERS, timeout=30)
        resp.raise_for_status()
        suffix = ".jpg"
        tmp = cache_dir / feed_id
        tmp.mkdir(parents=True, exist_ok=True)
        path = tmp / f"export_{index:02d}{suffix}"
        path.write_bytes(resp.content)
        return path
    except Exception:
        return None


def _make_thumbnail(src: Path, dest: Path) -> Path:
    with PILImage.open(src) as img:
        img = img.convert("RGB")
        img.thumbnail(THUMB_SIZE, PILImage.Resampling.LANCZOS)
        dest.parent.mkdir(parents=True, exist_ok=True)
        img.save(dest, format="JPEG", quality=85)
    return dest


def _max_images_in_rows(rows: list[dict[str, Any]]) -> int:
    counts = [len(row.get("image_urls") or []) for row in rows]
    if not counts:
        return 0
    return min(max(counts), MAX_IMAGES_CAP)


def build_excel_bytes(rows: list[dict[str, Any]]) -> bytes:
    rows = [dedupe_item_images(dict(row)) for row in rows]
    max_images = _max_images_in_rows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "小红书笔记"

    headers = [name for name, _ in TEXT_COLUMNS]
    for i in range(1, max_images + 1):
        headers.append(f"图片{i}")
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    cache_dir = Path(tempfile.gettempdir()) / "xhs_excel_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    thumb_dir = cache_dir / "thumbs"

    for row_idx, row in enumerate(rows, start=2):
        values = [row.get(key, "") for _, key in TEXT_COLUMNS]
        for _ in range(max_images):
            values.append("")
        ws.append(values)
        ws.row_dimensions[row_idx].height = 110

        for col_idx, (_, key) in enumerate(TEXT_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

        feed_id = row.get("feed_id") or f"row{row_idx}"
        urls = row.get("image_urls") or []
        local_paths = row.get("local_image_paths") or []

        for img_idx, url in enumerate(urls[:max_images], start=1):
            col = len(TEXT_COLUMNS) + img_idx
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = 22

            src_path: Path | None = None
            if img_idx <= len(local_paths) and local_paths[img_idx - 1]:
                candidate = Path(local_paths[img_idx - 1])
                if candidate.exists():
                    src_path = candidate
            if src_path is None:
                src_path = _load_image_path(url, feed_id, img_idx, cache_dir)

            if not src_path or not src_path.exists():
                continue

            thumb = thumb_dir / f"{feed_id}_{row_idx}_{img_idx}.jpg"
            try:
                _make_thumbnail(src_path, thumb)
                xl_img = XLImage(str(thumb))
                xl_img.width, xl_img.height = THUMB_SIZE
                ws.add_image(xl_img, f"{letter}{row_idx}")
            except Exception as exc:
                ws.cell(row=row_idx, column=col, value=f"[图片加载失败: {exc}]")

    for col_idx in range(1, len(TEXT_COLUMNS) + 1):
        letter = get_column_letter(col_idx)
        key = TEXT_COLUMNS[col_idx - 1][1]
        width = 18 if key in {"url", "desc", "video_script", "image_ocr_text"} else 12
        ws.column_dimensions[letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
